# -*- coding:utf-8 -*-
"""
==============================
Author:yuan
Time  :2021/4/23 15:33
file  :handler_excel.py
===============================
"""
import openpyxl
import os
from common.handle_path import TESTDATA_DIR


class HandlerExcel():
    '''excel表格操作方法'''

    def __init__(self, filename, sheetname):
        '''
        初始化操作文件及表单
        :param filename: 文件名
        :param sheetname: 表单名
        '''
        self.filename = filename
        self.sheetname = sheetname

    def read_excel_data(self):
        '''读取器excel测试用例数据'''
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        # 按行读取所有内容并转换成列表类型
        datas = list(sheet.rows)
        # 创建空列表，接收按行读取后的数据，每一行为一个列表添加到data_list列表中
        data_list = []
        for item in datas:
            # 临时空列表，每添加一个列表元素到data_list后重新生成空列表
            case_datas = []
            for data in item:
                case_datas.append(data.value)
            data_list.append(case_datas)
        # 表头为读取后的数据的第一行数据
        title = data_list[0]
        # 新建空列表，接收组装后的测试数据
        case_data = []
        for i in data_list[1:]:
            case_data.append(dict(zip(title, i)))
        return case_data

    def back_write_excel(self, row, column, value):
        '''
        回写测试数据到excel方法
        :param row:excel行
        :param column:excel列
        :param value:需要回写的值
        :return:None
        '''
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(row, column).value = value
        wb.save(self.filename)

if __name__ == '__main__':
    excel = HandlerExcel(os.path.join(TESTDATA_DIR,'case_data.xlsx'),'login')
    aa = excel.read_excel_data()
    print(aa)